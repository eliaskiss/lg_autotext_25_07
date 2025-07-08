from openpyxl import load_workbook
from datetime import datetime
from icecream import  ic
import time
from openpyxl.styles.builtins import styles

ic.configureOutput(includeContext=True)

file_name = 'public_bicycle.xlsx'

###############################################################
# Open Excel File
###############################################################
wb = load_workbook(file_name,
                   read_only=True, # read_only=True --> Lazy Loading 발생해서 대용량 엑셀파일 처리에 용이함
                   data_only=True) # True : Cell 함수의 결과값 (3), False: Cell 함수식(=A1+A2)

###############################################################
# Get Sheet List
###############################################################
ws_list = wb.sheetnames
# ic(ws_list)

###############################################################
# Select Worksheet
###############################################################
ws = wb.active # 현재 활성화 되어 있는 (저장하는순간의 워크시트) 시트를 선택
# ws = wb['대여소현황']
# ws = wb[ws_list[2]]
# ws = wb[wb.sheetnames[2]]

###############################################################
# Get Cell Value
###############################################################
# cell_a1 = ws['A1']
# ic(cell_a1)
# ic(type(cell_a1))
# ic(cell_a1.value)

###############################################################
# Get Formular Cell Value
###############################################################
# cell_total_lcd = ws['B2593']
# ic(cell_total_lcd.value)
# cell_total_qr = ws['B2594']
# ic(cell_total_qr.value)

###############################################################
# Get Datetime Cell Value
###############################################################
# cell_datetime = ws['G6']
# ic(type(cell_datetime.value))
# ic(cell_datetime.value)
# cell_datetime_value = cell_datetime.value.strftime('%Y-%m-%d %H:%M:%S')
# ic(cell_datetime_value)

###############################################################
# Get Datetime Cell Value
###############################################################
# cell_time = ws['B2596']
# ic(type(cell_time.value))
# ic(cell_time.value)
# cell_text_time = ws['B2597']
# ic(type(cell_text_time.value))
# ic(cell_text_time.value)

###############################################################
# Get Percent Cell Value
###############################################################
# cell_ratio = ws['B2600']
# ic(type(cell_ratio.value))
# ic(cell_ratio.value)
# cell_text_ratio = ws['B2599']
# ic(type(cell_text_ratio.value))
# ic(cell_text_ratio.value)

###############################################################
# Get Number Cell Value
###############################################################
# cell_number = ws['B2602']
# ic(type(cell_number.value))
# ic(cell_number.value)
# cell_text_number = ws['B2603']
# ic(type(cell_text_number.value))
# ic(cell_text_number.value)
# cell_float = ws['B2604']
# ic(type(cell_float.value))
# ic(cell_float.value)
# cell_with_text = ws['B2605']
# ic(type(cell_with_text.value))
# ic(cell_with_text.value)

###############################################################
# Get Cell Color
###############################################################
# fill_a6 = ws['A6'].fill
# ic(fill_a6.fgColor.index)
# ic(fill_a6.fgColor.type)
# if fill_a6.fgColor.type == 'rgb':
#     argb = fill_a6.fgColor.rgb
#     ic(argb)
#     argb = tuple(int(argb[i:i+2], 16) for i in range(0, len(argb), 2))
#     ic(argb)
# ic('----------------------------------------------------------')
# fill_b6 = ws['B6'].fill
# ic(fill_b6.fgColor.index)
# ic(fill_b6.fgColor.type)
# if fill_b6.fgColor.type == 'rgb':
#     argb = fill_b6.fgColor.rgb
#     ic(argb)
#     argb = tuple(int(argb[i:i+2], 16) for i in range(0, len(argb), 2))
#     ic(argb)
# ic('----------------------------------------------------------')

###############################################################
# Get Row Data
###############################################################
# for row in ws.rows:
#     row_values = [elem.value for elem in row]
#     ic(row_values)

# for index, row in enumerate(ws.rows):
#     if index >= 5 :
#         row_values = [elem.value for elem in row]
#         ic(row_values)
#         break

# for row in ws.iter_rows(min_row=6, max_row=2591):
#     row_values = [elem.value for elem in row]
#     ic(row_values)
#     break

###############################################################
# Get Data with slicing
###############################################################
# for row in ws['A6':'J20']:
#     row_values = [elem.value for elem in row]
#     ic(row_values)

###############################################################
# Get Hidden Row
###############################################################
# ic(ws.row_dimensions[6].hidden)
# ic(ws.row_dimensions[7].hidden)

###############################################################
# Get Hidden Column
###############################################################
# ic(ws.column_dimensions['F'].hidden)
# ic(ws.column_dimensions['G'].hidden)











